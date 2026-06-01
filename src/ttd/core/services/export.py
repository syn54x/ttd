"""Period CSV export — load, transform, and serialize."""

from __future__ import annotations

import csv
import io
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from uuid import UUID

from pydantic import BaseModel, ConfigDict

from ttd.core.domain.rates import effective_hourly_rate
from ttd.core.domain.rounding import effective_rounding_increment, round_hours_up
from ttd.core.exceptions import NotFoundError
from ttd.core.models.client import Client
from ttd.core.models.enums import BillingMode, EntryMode, enum_value
from ttd.core.models.project import Project
from ttd.core.models.time_entry import TimeEntry
from ttd.core.schemas import ExportPeriod
from ttd.core.services import clients as client_service
from ttd.core.services import projects as project_service
from ttd.core.services import time_entries as entry_service

_HOUR_SCALE = Decimal("0.01")
_MONEY_SCALE = Decimal("0.01")

_DETAIL_COLUMNS = (
    "row_type",
    "client",
    "project",
    "work_date",
    "entry_mode",
    "note",
    "time_from",
    "time_to",
    "billable",
    "hours",
    "currency",
    "rate",
    "amount",
)

_SUMMARY_COLUMNS = (
    "row_type",
    "level",
    "client",
    "project",
    "hours",
    "amount",
)


@dataclass(frozen=True, slots=True)
class LoadedEntry:
    """Entry with resolved client and project context."""

    client: Client
    project: Project
    entry: TimeEntry


class ExportDetailRow(BaseModel):
    model_config = ConfigDict(use_attribute_docstrings=True)

    client_name: str
    project_name: str
    work_date: date
    entry_mode: EntryMode
    note: str
    time_from: str
    time_to: str
    billable: bool
    hours: Decimal
    currency: str
    rate: Decimal | None
    amount: Decimal | None


class ExportSummaryRow(BaseModel):
    model_config = ConfigDict(use_attribute_docstrings=True)

    level: str
    client_name: str
    project_name: str
    hours: Decimal
    amount: Decimal | None


def _note_key(note: str | None) -> str:
    return (note or "").strip()


def _format_time(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%H:%M")


def _quantize_hours(value: Decimal) -> Decimal:
    return value.quantize(_HOUR_SCALE)


def _quantize_money(value: Decimal) -> Decimal:
    return value.quantize(_MONEY_SCALE)


def _rounded_hours_for_entry(loaded: LoadedEntry) -> Decimal:
    increment = effective_rounding_increment(loaded.client, loaded.project)
    return round_hours_up(loaded.entry.billable_hours, increment)


def _row_amount(
    loaded: LoadedEntry, *, rounded_hours: Decimal
) -> tuple[str, Decimal | None, Decimal | None]:
    if enum_value(loaded.project.billing_mode) != BillingMode.HOURLY.value:
        return "", None, None
    if not loaded.entry.billable:
        return "", None, None
    rate, currency = effective_hourly_rate(loaded.client, loaded.project)
    amount = _quantize_money(rounded_hours * rate)
    return currency, rate, amount


def _detail_row(
    loaded: LoadedEntry,
    *,
    hours: Decimal,
    note: str,
    time_from: str = "",
    time_to: str = "",
) -> ExportDetailRow:
    currency, rate, amount = _row_amount(loaded, rounded_hours=hours)
    return ExportDetailRow(
        client_name=loaded.client.name,
        project_name=loaded.project.name,
        work_date=loaded.entry.work_date,
        entry_mode=loaded.entry.entry_mode,
        note=note,
        time_from=time_from,
        time_to=time_to,
        billable=loaded.entry.billable,
        hours=_quantize_hours(hours),
        currency=currency,
        rate=rate,
        amount=amount,
    )


def _duration_rollup_row(
    template: LoadedEntry,
    *,
    hours: Decimal,
    note: str,
    billable: bool,
) -> ExportDetailRow:
    currency, rate, amount = _row_amount(template, rounded_hours=hours)
    if not billable:
        currency, rate, amount = "", None, None
    return ExportDetailRow(
        client_name=template.client.name,
        project_name=template.project.name,
        work_date=template.entry.work_date,
        entry_mode=EntryMode.DURATION,
        note=note,
        time_from="",
        time_to="",
        billable=billable,
        hours=_quantize_hours(hours),
        currency=currency,
        rate=rate,
        amount=amount,
    )


async def load_entries_for_export(
    period: ExportPeriod,
    *,
    client_id: UUID | None = None,
    project_id: UUID | None = None,
) -> list[LoadedEntry]:
    """Load ledger entries in range with optional client or project scope."""
    if project_id is not None:
        project = await project_service.get_project(project_id)
        client = await client_service.get_client(project.client_id)
        return _filter_loaded(
            [await _load_project_entries(client, project)],
            period,
        )

    if client_id is not None:
        client = await client_service.get_client(client_id)
        projects = await project_service.list_projects_for_client(client_id)
        loaded_groups = [
            await _load_project_entries(client, project) for project in projects
        ]
        return _filter_loaded(loaded_groups, period)

    loaded: list[LoadedEntry] = []
    for client in await client_service.list_clients():
        if client.id is None:
            raise NotFoundError("Client is missing an id")
        for project in await project_service.list_projects_for_client(client.id):
            loaded.extend(await _load_project_entries(client, project))
    return _filter_loaded([loaded], period)


async def _load_project_entries(client: Client, project: Project) -> list[LoadedEntry]:
    if project.id is None:
        raise NotFoundError("Project is missing an id")
    entries = await entry_service.list_time_entries_for_project(project.id)
    return [
        LoadedEntry(client=client, project=project, entry=entry) for entry in entries
    ]


def _filter_loaded(
    groups: list[list[LoadedEntry]], period: ExportPeriod
) -> list[LoadedEntry]:
    loaded: list[LoadedEntry] = []
    for group in groups:
        for item in group:
            if period.from_date <= item.entry.work_date <= period.to_date:
                loaded.append(item)
    loaded.sort(
        key=lambda item: (
            item.client.name,
            item.project.name,
            item.entry.work_date,
            enum_value(item.entry.entry_mode),
            _note_key(item.entry.note),
        )
    )
    return loaded


def _detail_row_values(row: ExportDetailRow) -> list[str]:
    return [
        "DETAIL",
        row.client_name,
        row.project_name,
        row.work_date.isoformat(),
        enum_value(row.entry_mode),
        row.note,
        row.time_from,
        row.time_to,
        "yes" if row.billable else "no",
        f"{row.hours:.2f}",
        row.currency,
        f"{row.rate:.2f}" if row.rate is not None else "",
        f"{row.amount:.2f}" if row.amount is not None else "",
    ]


def _summary_row_values(row: ExportSummaryRow) -> list[str]:
    return [
        "SUMMARY",
        row.level,
        row.client_name,
        row.project_name,
        f"{row.hours:.2f}",
        f"{row.amount:.2f}" if row.amount is not None else "",
    ]


async def _build_period_rows(
    period: ExportPeriod,
    *,
    client_id: UUID | None = None,
    project_id: UUID | None = None,
) -> tuple[list[ExportDetailRow], list[ExportSummaryRow]]:
    loaded = await load_entries_for_export(
        period, client_id=client_id, project_id=project_id
    )
    detail_rows = build_detail_rows(loaded)
    summary_rows = build_summary_rows(detail_rows)
    return detail_rows, summary_rows


async def export_period_csv(
    period: ExportPeriod,
    *,
    client_id: UUID | None = None,
    project_id: UUID | None = None,
) -> str:
    """Load, transform, and render a period CSV export."""
    detail_rows, summary_rows = await _build_period_rows(
        period, client_id=client_id, project_id=project_id
    )
    return render_period_csv(detail_rows, summary_rows)


async def export_period_xlsx(
    period: ExportPeriod,
    *,
    client_id: UUID | None = None,
    project_id: UUID | None = None,
) -> bytes:
    """Load, transform, and render a period XLSX export."""
    detail_rows, summary_rows = await _build_period_rows(
        period, client_id=client_id, project_id=project_id
    )
    return render_period_xlsx(detail_rows, summary_rows)


async def export_period_numbers(
    period: ExportPeriod,
    *,
    client_id: UUID | None = None,
    project_id: UUID | None = None,
) -> bytes:
    """Load, transform, and render a period Numbers export."""
    detail_rows, summary_rows = await _build_period_rows(
        period, client_id=client_id, project_id=project_id
    )
    return render_period_numbers(detail_rows, summary_rows)


def build_detail_rows(loaded_entries: list[LoadedEntry]) -> list[ExportDetailRow]:
    """Apply rounding, duration rollup, and hourly amounts."""
    interval_rows: list[ExportDetailRow] = []
    duration_buckets: dict[tuple[UUID, date, str], list[LoadedEntry]] = defaultdict(
        list
    )

    for loaded in loaded_entries:
        mode = enum_value(loaded.entry.entry_mode)
        if mode == EntryMode.INTERVAL.value:
            rounded = _rounded_hours_for_entry(loaded)
            interval_rows.append(
                _detail_row(
                    loaded,
                    hours=rounded,
                    note=_note_key(loaded.entry.note),
                    time_from=_format_time(loaded.entry.started_at),
                    time_to=_format_time(loaded.entry.ended_at),
                )
            )
            continue

        if loaded.project.id is None:
            raise NotFoundError("Project is missing an id")
        key = (loaded.project.id, loaded.entry.work_date, _note_key(loaded.entry.note))
        duration_buckets[key].append(loaded)

    duration_rows: list[ExportDetailRow] = []

    def _bucket_sort_key(item: tuple[UUID, date, str]) -> tuple[date, str, str]:
        return (item[1], item[2], str(item[0]))

    for key in sorted(duration_buckets, key=_bucket_sort_key):
        entries = duration_buckets[key]
        total_hours = Decimal(sum(_rounded_hours_for_entry(item) for item in entries))
        billable = any(item.entry.billable for item in entries)
        row = _duration_rollup_row(
            entries[0],
            hours=total_hours,
            note=key[2],
            billable=billable,
        )
        project_mode = enum_value(entries[0].project.billing_mode)
        if billable and project_mode == BillingMode.HOURLY.value:
            amount = Decimal("0")
            for item in entries:
                if not item.entry.billable:
                    continue
                _, _, entry_amount = _row_amount(
                    item, rounded_hours=_rounded_hours_for_entry(item)
                )
                if entry_amount is not None:
                    amount += entry_amount
            quantized = _quantize_money(amount) if amount else None
            row = row.model_copy(update={"amount": quantized})
        duration_rows.append(row)

    return sorted(
        interval_rows + duration_rows,
        key=lambda row: (
            row.client_name,
            row.project_name,
            row.work_date,
            enum_value(row.entry_mode),
            row.note,
        ),
    )


def build_summary_rows(detail_rows: list[ExportDetailRow]) -> list[ExportSummaryRow]:
    """Aggregate billable hours and hourly dollars by project and client."""
    if not detail_rows:
        return []

    project_hours: dict[tuple[str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    project_amounts: dict[tuple[str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    client_hours: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    client_amounts: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))

    for row in detail_rows:
        if not row.billable:
            continue
        project_key = (row.client_name, row.project_name)
        project_hours[project_key] += row.hours
        client_hours[row.client_name] += row.hours
        if row.amount is not None:
            project_amounts[project_key] += row.amount
            client_amounts[row.client_name] += row.amount

    summary: list[ExportSummaryRow] = []
    for (client_name, project_name), hours in sorted(project_hours.items()):
        amount = project_amounts.get((client_name, project_name))
        summary.append(
            ExportSummaryRow(
                level="project",
                client_name=client_name,
                project_name=project_name,
                hours=_quantize_hours(hours),
                amount=_quantize_money(amount) if amount is not None else None,
            )
        )

    for client_name, hours in sorted(client_hours.items()):
        amount = client_amounts.get(client_name)
        summary.append(
            ExportSummaryRow(
                level="client",
                client_name=client_name,
                project_name="",
                hours=_quantize_hours(hours),
                amount=_quantize_money(amount) if amount is not None else None,
            )
        )

    return summary


def render_period_csv(
    detail_rows: list[ExportDetailRow],
    summary_rows: list[ExportSummaryRow],
) -> str:
    """Serialize detail and summary blocks to CSV text."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(_DETAIL_COLUMNS)
    for row in detail_rows:
        writer.writerow(_detail_row_values(row))

    if summary_rows:
        writer.writerow([])
        writer.writerow(_SUMMARY_COLUMNS)
        for row in summary_rows:
            writer.writerow(_summary_row_values(row))

    return buffer.getvalue()


def _format_xlsx_sheet(
    sheet,
    *,
    table_name: str,
    num_columns: int,
    row_count: int,
) -> None:
    """Bold/freeze row 1 and register an unstyled Excel table for header detection."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table

    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    if row_count < 1:
        return

    last_col = get_column_letter(num_columns)
    sheet.add_table(
        Table(
            displayName=table_name,
            ref=f"A1:{last_col}{row_count}",
            headerRowCount=1,
        )
    )


def render_period_xlsx(
    detail_rows: list[ExportDetailRow],
    summary_rows: list[ExportSummaryRow],
) -> bytes:
    """Serialize detail and summary blocks to XLSX (Log + Summary sheets)."""
    from openpyxl import Workbook

    workbook = Workbook()
    log_sheet = workbook.active
    assert log_sheet is not None
    log_sheet.title = "Log"
    log_sheet.append(list(_DETAIL_COLUMNS))
    for row in detail_rows:
        log_sheet.append(_detail_row_values(row))
    _format_xlsx_sheet(
        log_sheet,
        table_name="TtdLog",
        num_columns=len(_DETAIL_COLUMNS),
        row_count=1 + len(detail_rows),
    )

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(list(_SUMMARY_COLUMNS))
    for row in summary_rows:
        summary_sheet.append(_summary_row_values(row))
    _format_xlsx_sheet(
        summary_sheet,
        table_name="TtdSummary",
        num_columns=len(_SUMMARY_COLUMNS),
        row_count=1 + len(summary_rows),
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_numbers_table(
    table,
    columns: tuple[str, ...],
    cell_rows: list[list[str]],
) -> None:
    """Write column headings and body rows with one Numbers header row."""
    table.num_header_rows = 1
    table.num_header_cols = 0
    for col_idx, heading in enumerate(columns):
        table.write(0, col_idx, heading)
    for row_idx, values in enumerate(cell_rows, start=1):
        for col_idx, value in enumerate(values):
            table.write(row_idx, col_idx, value)


def render_period_numbers(
    detail_rows: list[ExportDetailRow],
    summary_rows: list[ExportSummaryRow],
) -> bytes:
    """Serialize detail and summary blocks to Numbers (Log + Summary sheets)."""
    import tempfile
    from pathlib import Path

    from numbers_parser import Document

    document = Document()
    log_sheet = document.sheets[0]
    log_sheet.name = "Log"
    log_table = log_sheet.tables[0]
    log_table.name = "Log"
    _write_numbers_table(
        log_table,
        _DETAIL_COLUMNS,
        [_detail_row_values(row) for row in detail_rows],
    )

    document.add_sheet("Summary")
    summary_sheet = document.sheets[1]
    summary_table = summary_sheet.tables[0]
    summary_table.name = "Summary"
    _write_numbers_table(
        summary_table,
        _SUMMARY_COLUMNS,
        [_summary_row_values(row) for row in summary_rows],
    )

    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "period.numbers"
        document.save(path)
        return path.read_bytes()
