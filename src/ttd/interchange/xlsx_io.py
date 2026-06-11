from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from ttd.core.errors import TtdError
from ttd.interchange.base import Format, register
from ttd.interchange.model import COLUMNS, EntryRecord


def write_xlsx(records: list[EntryRecord], path: Path, meta: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Entries"
    ws.append(COLUMNS)
    ws.freeze_panes = "A2"
    for record in records:
        ws.append(
            [
                record.uid,
                record.client,
                record.project,
                record.date,  # real date cell
                record.start,  # real time cells
                record.end,
                float(record.hours),
                record.seconds,
                record.note,
                record.tags,
                record.billable,
                record.invoice_number,
            ]
        )
    for col, width in zip(ws.columns, (36, 14, 16, 12, 10, 10, 8, 9, 30, 14, 9, 14), strict=False):
        ws.column_dimensions[col[0].column_letter].width = width
    wb.save(path)


def read_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # zipfile.BadZipFile and friends
        raise TtdError(f"{path} is not a readable xlsx file: {exc}") from exc
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip().lower() if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    out = [
        dict(zip(header, row, strict=False))
        for row in rows
        if any(v is not None and str(v).strip() for v in row)
    ]
    wb.close()
    return out


register(Format("xlsx", ("xlsx",), write_xlsx, read_xlsx))
