from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal

import pytest
from pydantic import ValidationError as PydanticValidationError

from ttd.core.exceptions import NotFoundError
from ttd.core.models.enums import BillingMode, EntryMode
from ttd.core.schemas import (
    CreateClient,
    CreateDurationEntry,
    CreateIntervalEntry,
    CreateProject,
    ExportPeriod,
)
from ttd.core.services import clients as client_service
from ttd.core.services import projects as project_service
from ttd.core.services import time_entries as entry_service
from ttd.core.services.export import (
    build_detail_rows,
    build_summary_rows,
    export_period_csv,
    export_period_numbers,
    export_period_xlsx,
    load_entries_for_export,
)


async def _hourly_setup(
    db,
    *,
    client_rounding: int | None = None,
    project_rounding: int | None = None,
) -> tuple:
    client = await client_service.create_client(
        CreateClient(
            name="Acme",
            default_hourly_rate=Decimal("150"),
            currency="USD",
            rounding_increment_minutes=client_rounding,
        )
    )
    project = await project_service.create_project(
        CreateProject(
            client_id=client.id,
            name="Website",
            billing_mode=BillingMode.HOURLY,
            rounding_increment_minutes=project_rounding,
        )
    )
    return client, project


async def test_export_period_rejects_inverted_dates() -> None:
    with pytest.raises(PydanticValidationError):
        ExportPeriod(from_date=date(2026, 5, 31), to_date=date(2026, 5, 1))


async def test_load_entries_respects_period_bounds(db) -> None:
    _, project = await _hourly_setup(db)
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project.id,
            work_date=date(2026, 5, 1),
            billable_hours=Decimal("1"),
        )
    )
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project.id,
            work_date=date(2026, 6, 1),
            billable_hours=Decimal("2"),
        )
    )
    loaded = await load_entries_for_export(
        ExportPeriod(from_date=date(2026, 5, 1), to_date=date(2026, 5, 31))
    )
    assert len(loaded) == 1
    assert loaded[0].entry.billable_hours == Decimal("1")


async def test_load_entries_client_filter(db) -> None:
    client_a, project_a = await _hourly_setup(db)
    client_b = await client_service.create_client(
        CreateClient(name="Beta", default_hourly_rate=Decimal("100"), currency="USD")
    )
    project_b = await project_service.create_project(
        CreateProject(
            client_id=client_b.id,
            name="Other",
            billing_mode=BillingMode.HOURLY,
        )
    )
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project_a.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("1"),
        )
    )
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project_b.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("2"),
        )
    )
    loaded = await load_entries_for_export(
        ExportPeriod(from_date=date(2026, 5, 1), to_date=date(2026, 5, 31)),
        client_id=client_a.id,
    )
    assert len(loaded) == 1
    assert loaded[0].client.name == "Acme"


async def test_load_entries_unknown_client(db) -> None:
    from uuid import uuid4

    with pytest.raises(NotFoundError):
        await load_entries_for_export(
            ExportPeriod(from_date=date(2026, 5, 1), to_date=date(2026, 5, 31)),
            client_id=uuid4(),
        )


async def test_duration_rollup_and_interval_rows(db) -> None:
    _, project = await _hourly_setup(db)
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("1"),
            note="Standup",
        )
    )
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("2"),
            note="Standup",
        )
    )
    await entry_service.create_interval_entry(
        CreateIntervalEntry(
            project_id=project.id,
            work_date=date(2026, 5, 10),
            started_at=datetime(2026, 5, 10, 9, 0, tzinfo=UTC),
            ended_at=datetime(2026, 5, 10, 10, 0, tzinfo=UTC),
        )
    )
    await entry_service.create_interval_entry(
        CreateIntervalEntry(
            project_id=project.id,
            work_date=date(2026, 5, 10),
            started_at=datetime(2026, 5, 10, 11, 0, tzinfo=UTC),
            ended_at=datetime(2026, 5, 10, 12, 0, tzinfo=UTC),
        )
    )
    loaded = await load_entries_for_export(
        ExportPeriod(from_date=date(2026, 5, 1), to_date=date(2026, 5, 31))
    )
    detail = build_detail_rows(loaded)
    duration_rows = [row for row in detail if row.entry_mode == EntryMode.DURATION]
    interval_rows = [row for row in detail if row.entry_mode == EntryMode.INTERVAL]
    assert len(duration_rows) == 1
    assert duration_rows[0].hours == Decimal("3.00")
    assert len(interval_rows) == 2


async def test_different_notes_split_duration_rows(db) -> None:
    _, project = await _hourly_setup(db)
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("1"),
            note="API",
        )
    )
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("2"),
            note=None,
        )
    )
    loaded = await load_entries_for_export(
        ExportPeriod(from_date=date(2026, 5, 1), to_date=date(2026, 5, 31))
    )
    detail = build_detail_rows(loaded)
    assert len(detail) == 2
    notes = {row.note for row in detail}
    assert notes == {"API", ""}


async def test_rounding_at_export_only(db) -> None:
    _, project = await _hourly_setup(db, client_rounding=15)
    entry = await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("2.10"),
        )
    )
    loaded = await load_entries_for_export(
        ExportPeriod(from_date=date(2026, 5, 1), to_date=date(2026, 5, 31))
    )
    detail = build_detail_rows(loaded)
    assert detail[0].hours == Decimal("2.25")
    stored = await entry_service.get_time_entry(entry.id)
    assert stored.billable_hours == Decimal("2.10")


async def test_hourly_and_fixed_price_columns(db, sample_client) -> None:
    hourly = await project_service.create_project(
        CreateProject(
            client_id=sample_client.id,
            name="Hourly",
            billing_mode=BillingMode.HOURLY,
        )
    )
    fixed = await project_service.create_project(
        CreateProject(
            client_id=sample_client.id,
            name="Fixed",
            billing_mode=BillingMode.FIXED_PRICE,
            contract_total=Decimal("5000"),
            currency="USD",
        )
    )
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=hourly.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("10"),
        )
    )
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=fixed.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("5"),
        )
    )
    loaded = await load_entries_for_export(
        ExportPeriod(from_date=date(2026, 5, 1), to_date=date(2026, 5, 31))
    )
    detail = build_detail_rows(loaded)
    hourly_row = next(row for row in detail if row.project_name == "Hourly")
    fixed_row = next(row for row in detail if row.project_name == "Fixed")
    assert hourly_row.amount == Decimal("1500.00")
    assert fixed_row.amount is None
    assert fixed_row.currency == ""


async def test_non_billable_excluded_from_summary_dollars(db) -> None:
    _, project = await _hourly_setup(db)
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("2"),
            billable=False,
        )
    )
    loaded = await load_entries_for_export(
        ExportPeriod(from_date=date(2026, 5, 1), to_date=date(2026, 5, 31))
    )
    detail = build_detail_rows(loaded)
    assert detail[0].billable is False
    assert detail[0].amount is None
    summary = build_summary_rows(detail)
    assert summary == []


async def test_csv_contains_detail_and_summary(db) -> None:
    _, project = await _hourly_setup(db)
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("2"),
        )
    )
    csv_text = await export_period_csv(
        ExportPeriod(from_date=date(2026, 5, 1), to_date=date(2026, 5, 31))
    )
    assert "row_type,client,project" in csv_text
    assert "DETAIL,Acme,Website" in csv_text
    assert "SUMMARY,project,Acme,Website" in csv_text
    assert "SUMMARY,client,Acme," in csv_text


async def test_reexport_reflects_edits_without_mutating_stored_hours(db) -> None:
    _, project = await _hourly_setup(db)
    entry = await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("2"),
        )
    )
    period = ExportPeriod(from_date=date(2026, 5, 1), to_date=date(2026, 5, 31))
    first = await export_period_csv(period)
    assert "2.00" in first
    from ttd.core.schemas import UpdateDurationEntry

    await entry_service.update_duration_entry(
        entry.id,
        UpdateDurationEntry(billable_hours=Decimal("3")),
    )
    second = await export_period_csv(period)
    assert "3.00" in second
    stored = await entry_service.get_time_entry(entry.id)
    assert stored.billable_hours == Decimal("3")


async def test_empty_period_csv_has_header_only(db) -> None:
    csv_text = await export_period_csv(
        ExportPeriod(from_date=date(2026, 5, 1), to_date=date(2026, 5, 31))
    )
    lines = [line for line in csv_text.strip().splitlines() if line]
    assert len(lines) == 1
    assert lines[0].startswith("row_type,")


async def test_xlsx_has_log_and_summary_sheets(db) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    _, project = await _hourly_setup(db)
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("2"),
        )
    )
    xlsx_bytes = await export_period_xlsx(
        ExportPeriod(from_date=date(2026, 5, 1), to_date=date(2026, 5, 31))
    )
    workbook = load_workbook(BytesIO(xlsx_bytes), read_only=True)
    assert workbook.sheetnames == ["Log", "Summary"]

    log = workbook["Log"]
    log_rows = list(log.iter_rows(values_only=True))
    assert log_rows[0][0] == "row_type"
    assert log_rows[1][0] == "DETAIL"
    assert log_rows[1][1] == "Acme"

    summary = workbook["Summary"]
    summary_rows = list(summary.iter_rows(values_only=True))
    assert summary_rows[0] == (
        "row_type",
        "level",
        "client",
        "project",
        "hours",
        "amount",
    )
    assert summary_rows[1][0] == "SUMMARY"
    assert summary_rows[1][1] == "project"
    workbook.close()

    workbook = load_workbook(BytesIO(xlsx_bytes))
    assert workbook["Log"].freeze_panes == "A2"
    assert workbook["Summary"].freeze_panes == "A2"
    log_table = workbook["Log"].tables["TtdLog"]
    summary_table = workbook["Summary"].tables["TtdSummary"]
    assert log_table.headerRowCount == 1
    assert summary_table.headerRowCount == 1
    assert log_table.tableStyleInfo is None
    assert summary_table.tableStyleInfo is None
    assert log_table.ref == "A1:M2"
    assert all(cell.font.bold for cell in workbook["Log"][1])
    assert all(cell.font.bold for cell in workbook["Summary"][1])


async def test_numbers_has_log_and_summary_sheets(db) -> None:
    import tempfile
    from pathlib import Path

    from numbers_parser import Document

    _, project = await _hourly_setup(db)
    await entry_service.create_duration_entry(
        CreateDurationEntry(
            project_id=project.id,
            work_date=date(2026, 5, 10),
            billable_hours=Decimal("2"),
        )
    )
    numbers_bytes = await export_period_numbers(
        ExportPeriod(from_date=date(2026, 5, 1), to_date=date(2026, 5, 31))
    )
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "period.numbers"
        path.write_bytes(numbers_bytes)
        document = Document(path)

    assert [sheet.name for sheet in document.sheets] == ["Log", "Summary"]

    log = document.sheets[0].tables[0]
    assert log.name == "Log"
    assert log.num_header_rows == 1
    assert log.num_header_cols == 0
    assert log.cell(0, 0).value == "row_type"
    assert log.cell(1, 0).value == "DETAIL"
    assert log.cell(1, 1).value == "Acme"
    assert log.cell(1, 9).value == "2.00"

    summary = document.sheets[1].tables[0]
    assert summary.name == "Summary"
    assert summary.num_header_rows == 1
    assert summary.cell(0, 0).value == "row_type"
    assert summary.cell(1, 0).value == "SUMMARY"
    assert summary.cell(1, 1).value == "project"
