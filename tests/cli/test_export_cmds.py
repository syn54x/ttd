from __future__ import annotations

import asyncio
from pathlib import Path

import pytest

from ttd.cli.interactive import set_invocation_tokens
from ttd.cli.main import app
from ttd.core.db import close_db, init_db


@pytest.fixture
def cli_settings(monkeypatch, settings):
    monkeypatch.setattr("ttd.core.config.get_settings", lambda: settings)
    return settings


@pytest.fixture
def cli_db(cli_settings, reset_db_state):
    asyncio.run(init_db(cli_settings))
    yield cli_settings
    asyncio.run(close_db())


def run_cli(argv: list[str]) -> None:
    set_invocation_tokens(argv)
    with pytest.raises(SystemExit) as exc_info:
        app(argv)
    assert exc_info.value.code == 0


def test_export_to_stdout(cli_db, capsys) -> None:
    run_cli(["client", "add", "Acme", "--rate", "150"])
    run_cli(["project", "add", "--client", "Acme", "--name", "Website"])
    run_cli(
        [
            "log",
            "--client",
            "Acme",
            "--project",
            "Website",
            "--date",
            "2026-05-20",
            "--hours",
            "2.5",
            "--note",
            "API work",
        ]
    )
    run_cli(["export", "--from", "2026-05-01", "--to", "2026-05-31"])
    out = capsys.readouterr().out
    assert "DETAIL,Acme,Website" in out
    assert "API work" in out


def test_export_to_file(cli_db, tmp_path: Path) -> None:
    run_cli(["client", "add", "Acme", "--rate", "150"])
    run_cli(["project", "add", "--client", "Acme", "--name", "Website"])
    run_cli(
        [
            "log",
            "--client",
            "Acme",
            "--project",
            "Website",
            "--date",
            "2026-05-20",
            "--hours",
            "1",
        ]
    )
    output = tmp_path / "period.csv"
    run_cli(
        [
            "export",
            "--from",
            "2026-05-01",
            "--to",
            "2026-05-31",
            "--output",
            str(output),
        ]
    )
    text = output.read_text(encoding="utf-8")
    assert "DETAIL,Acme,Website" in text


def test_export_to_xlsx(cli_db, tmp_path: Path) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    run_cli(["client", "add", "Acme", "--rate", "150"])
    run_cli(["project", "add", "--client", "Acme", "--name", "Website"])
    run_cli(
        [
            "log",
            "--client",
            "Acme",
            "--project",
            "Website",
            "--date",
            "2026-05-20",
            "--hours",
            "1",
        ]
    )
    output = tmp_path / "period.xlsx"
    run_cli(
        [
            "export",
            "--from",
            "2026-05-01",
            "--to",
            "2026-05-31",
            "--output",
            str(output),
        ]
    )
    workbook = load_workbook(BytesIO(output.read_bytes()), read_only=True)
    assert "Log" in workbook.sheetnames
    assert "Summary" in workbook.sheetnames


def test_export_to_numbers(cli_db, tmp_path: Path) -> None:
    from numbers_parser import Document

    run_cli(["client", "add", "Acme", "--rate", "150"])
    run_cli(["project", "add", "--client", "Acme", "--name", "Website"])
    run_cli(
        [
            "log",
            "--client",
            "Acme",
            "--project",
            "Website",
            "--date",
            "2026-05-20",
            "--hours",
            "1",
        ]
    )
    output = tmp_path / "period.numbers"
    run_cli(
        [
            "export",
            "--from",
            "2026-05-01",
            "--to",
            "2026-05-31",
            "--output",
            str(output),
        ]
    )
    document = Document(output)
    assert document.sheets[0].name == "Log"
    assert document.sheets[1].name == "Summary"
    log = document.sheets[0].tables[0]
    assert log.num_header_rows == 1
    assert log.cell(1, 0).value == "DETAIL"
